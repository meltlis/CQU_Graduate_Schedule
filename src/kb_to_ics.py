#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from icalendar import Calendar, Event
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

FIELD_ORDER = ("班号", "代码", "课程", "教师", "周次", "节次", "教室")

# User-provided period time mapping.
PERIOD_TIME_MAP: dict[tuple[int, ...], tuple[time, time]] = {
    (1, 2): (time(8, 30), time(10, 10)),
    (3, 4): (time(10, 30), time(12, 10)),
    (6, 7): (time(14, 25), time(16, 5)),
    (8, 9): (time(16, 25), time(18, 5)),
}

MERGE_PERIOD_RULES: tuple[tuple[tuple[int, ...], tuple[int, ...], str], ...] = (
    ((1, 2), (3, 4), "第1-4节"),
    ((6, 7), (8, 9), "第6-9节"),
)

PERIOD_LABEL_MAP: dict[tuple[int, ...], str] = {
    (1, 2): "第1,2节",
    (3, 4): "第3,4节",
    (6, 7): "第6,7节",
    (8, 9): "第8,9节",
}


@dataclass(frozen=True)
class CourseRecord:
    class_no: str
    course_code: str
    course_name: str
    teacher_name: str
    week_expr: str
    period_expr: str
    classroom: str
    weekday_name: str
    weekday_offset: int
    source_cell: str


@dataclass(frozen=True)
class EventOccurrence:
    record: CourseRecord
    week_no: int
    event_date: date
    period_key: tuple[int, ...]
    period_label: str
    start_time: time
    end_time: time


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_labeled_text(text: str) -> tuple[str, str]:
    match = re.match(r"^\s*([^：:]+)\s*[：:]\s*(.*)\s*$", text)
    if not match:
        return "", text.strip()
    return match.group(1).strip(), match.group(2).strip()


def parse_week_expression(week_expr: str) -> list[int]:
    cleaned = week_expr.replace("，", ",").replace("、", ",").replace(" ", "")
    if not cleaned:
        raise ValueError("Empty week expression")

    weeks: list[int] = []
    for part in cleaned.split(","):
        if not part:
            continue
        if "-" in part:
            start_text, end_text = part.split("-", 1)
            start = int(start_text)
            end = int(end_text)
            if start > end:
                raise ValueError(f"Invalid week range: {part}")
            weeks.extend(range(start, end + 1))
        else:
            weeks.append(int(part))

    unique_weeks: list[int] = []
    seen: set[int] = set()
    for week in weeks:
        if week not in seen:
            unique_weeks.append(week)
            seen.add(week)
    return unique_weeks


def parse_period_expression(period_expr: str) -> tuple[int, ...]:
    numbers = tuple(int(x) for x in re.findall(r"\d+", period_expr))
    if not numbers:
        raise ValueError(f"Invalid period expression: {period_expr}")
    if numbers not in PERIOD_TIME_MAP:
        supported = ", ".join(f"{k}" for k in PERIOD_TIME_MAP)
        raise ValueError(f"Unsupported period '{period_expr}'. Supported: {supported}")
    return numbers


def extract_course_records(sheet: Worksheet) -> list[CourseRecord]:
    records: list[CourseRecord] = []
    for row in range(1, sheet.max_row + 1):
        for col in range(2, 9):  # B..H = Monday..Sunday
            first_cell = normalize_text(sheet.cell(row=row, column=col).value)
            if not first_cell.startswith("班号"):
                continue

            values: dict[str, str] = {}
            for idx, expected_label in enumerate(FIELD_ORDER):
                current_row = row + idx
                current_ref = f"{get_column_letter(col)}{current_row}"
                raw = normalize_text(sheet.cell(row=current_row, column=col).value)
                if not raw:
                    if expected_label == "教室":
                        values[expected_label] = ""
                        continue
                    raise ValueError(f"Missing '{expected_label}' at {current_ref}")

                actual_label, actual_value = parse_labeled_text(raw)
                if actual_label and actual_label != expected_label:
                    raise ValueError(
                        f"Unexpected label at {current_ref}: expected '{expected_label}', got '{actual_label}'"
                    )
                values[expected_label] = actual_value

            weekday_name = normalize_text(sheet.cell(row=1, column=col).value) or f"周{col - 1}"
            source_cell = f"{get_column_letter(col)}{row}"
            records.append(
                CourseRecord(
                    class_no=values["班号"],
                    course_code=values["代码"],
                    course_name=values["课程"],
                    teacher_name=values["教师"],
                    week_expr=values["周次"],
                    period_expr=values["节次"],
                    classroom=values["教室"],
                    weekday_name=weekday_name,
                    weekday_offset=col - 2,
                    source_cell=source_cell,
                )
            )
    return records


def build_uid(record: CourseRecord, event_date: date, start_time: time, end_time: time) -> str:
    seed = "|".join(
        [
            record.class_no,
            record.course_code,
            record.course_name,
            record.teacher_name,
            record.classroom,
            record.weekday_name,
            event_date.isoformat(),
            start_time.isoformat(),
            end_time.isoformat(),
            record.source_cell,
        ]
    )
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()
    return f"{digest}@kb1-schedule-ics.local"


def build_description(record: CourseRecord, week_no: int, event_date: date, period_label: str) -> str:
    display_room = record.classroom if record.classroom else "未安排教室"
    first_line = f"{record.course_name} | {display_room} | {record.teacher_name}"
    lines = [
        first_line,
        f"班号：{record.class_no}",
        f"代码：{record.course_code}",
        f"课程：{record.course_name}",
        f"教师：{record.teacher_name}",
        f"教室：{record.classroom}",
        f"周次：{record.week_expr}",
        f"节次：{period_label}",
        f"星期：{record.weekday_name}",
        f"第{week_no}周",
        f"日期：{event_date.isoformat()}",
    ]
    return "\n".join(lines)


def build_summary(record: CourseRecord) -> str:
    display_room = record.classroom if record.classroom else "未安排教室"
    return f"{record.course_name}({display_room})"


def expand_occurrences(records: list[CourseRecord], semester_monday: date) -> list[EventOccurrence]:
    occurrences: list[EventOccurrence] = []
    for record in records:
        weeks = parse_week_expression(record.week_expr)
        period_key = parse_period_expression(record.period_expr)
        start_time, end_time = PERIOD_TIME_MAP[period_key]
        period_label = PERIOD_LABEL_MAP.get(period_key, record.period_expr)

        for week_no in weeks:
            event_date = semester_monday + timedelta(days=(week_no - 1) * 7 + record.weekday_offset)
            occurrences.append(
                EventOccurrence(
                    record=record,
                    week_no=week_no,
                    event_date=event_date,
                    period_key=period_key,
                    period_label=period_label,
                    start_time=start_time,
                    end_time=end_time,
                )
            )
    return occurrences


def merge_occurrences(occurrences: list[EventOccurrence]) -> list[EventOccurrence]:
    grouped: dict[tuple[object, ...], dict[tuple[int, ...], EventOccurrence]] = {}
    for occ in occurrences:
        key = (
            occ.event_date,
            occ.week_no,
            occ.record.class_no,
            occ.record.course_code,
            occ.record.course_name,
            occ.record.teacher_name,
            occ.record.classroom,
            occ.record.weekday_name,
        )
        grouped.setdefault(key, {})[occ.period_key] = occ

    merged_results: list[EventOccurrence] = []
    for period_map in grouped.values():
        consumed: set[tuple[int, ...]] = set()
        for left_key, right_key, merged_label in MERGE_PERIOD_RULES:
            left = period_map.get(left_key)
            right = period_map.get(right_key)
            if left is None or right is None:
                continue
            start_time = min(left.start_time, right.start_time)
            end_time = max(left.end_time, right.end_time)
            merged_results.append(
                EventOccurrence(
                    record=left.record,
                    week_no=left.week_no,
                    event_date=left.event_date,
                    period_key=left_key + right_key,
                    period_label=merged_label,
                    start_time=start_time,
                    end_time=end_time,
                )
            )
            consumed.add(left_key)
            consumed.add(right_key)

        for period_key, occ in period_map.items():
            if period_key in consumed:
                continue
            merged_results.append(occ)

    return merged_results


def build_calendar(
    records: list[CourseRecord],
    semester_monday: date,
    tz_name: str,
    calendar_name: str,
) -> tuple[Calendar, int]:
    tz = ZoneInfo(tz_name)
    dtstamp = datetime.now(timezone.utc)

    calendar = Calendar()
    calendar.add("prodid", "-//kb1 schedule to ics//CN")
    calendar.add("version", "2.0")
    calendar.add("calscale", "GREGORIAN")
    calendar.add("X-WR-CALNAME", calendar_name)
    calendar.add("X-WR-TIMEZONE", tz_name)

    sorted_events: list[tuple[datetime, Event]] = []
    occurrences = merge_occurrences(expand_occurrences(records, semester_monday))
    for occ in occurrences:
        record = occ.record
        start_dt = datetime.combine(occ.event_date, occ.start_time, tzinfo=tz)
        end_dt = datetime.combine(occ.event_date, occ.end_time, tzinfo=tz)

        event = Event()
        event.add("uid", build_uid(record, occ.event_date, occ.start_time, occ.end_time))
        event.add("dtstamp", dtstamp)
        event.add("dtstart", start_dt)
        event.add("dtend", end_dt)
        event.add("summary", build_summary(record))
        if record.classroom:
            event.add("location", record.classroom)
        event.add("description", build_description(record, occ.week_no, occ.event_date, occ.period_label))

        sorted_events.append((start_dt, event))

    sorted_events.sort(key=lambda item: item[0])
    for _, event in sorted_events:
        calendar.add_component(event)

    return calendar, len(sorted_events)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert kb.xlsx schedule to ICS calendar.")
    parser.add_argument("--input", default="kb.xlsx", help="Input xlsx file path.")
    parser.add_argument("--output", default="kb.ics", help="Output ics file path.")
    parser.add_argument(
        "--semester-monday",
        default="2026-03-02",
        help="The first Monday of the semester, format YYYY-MM-DD.",
    )
    parser.add_argument(
        "--timezone",
        default="Asia/Shanghai",
        help="IANA timezone name used for calendar events.",
    )
    parser.add_argument(
        "--calendar-name",
        default="2026春季课表",
        help="Calendar name shown by calendar clients.",
    )
    parser.add_argument(
        "--sheet",
        default="",
        help="Worksheet name. Defaults to the first worksheet.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    semester_monday = date.fromisoformat(args.semester_monday)
    if semester_monday.weekday() != 0:
        raise ValueError(f"semester-monday must be a Monday, got {semester_monday.isoformat()}")

    workbook = load_workbook(filename=input_path, data_only=True)
    if args.sheet:
        if args.sheet not in workbook.sheetnames:
            raise ValueError(f"Worksheet '{args.sheet}' not found. Available: {workbook.sheetnames}")
        sheet = workbook[args.sheet]
    else:
        sheet = workbook.worksheets[0]

    records = extract_course_records(sheet)
    if not records:
        raise RuntimeError("No course records found in worksheet.")

    calendar, event_count = build_calendar(records, semester_monday, args.timezone, args.calendar_name)
    output_path = Path(args.output)
    output_path.write_bytes(calendar.to_ical())

    print(f"Parsed {len(records)} course records from '{sheet.title}'.")
    print(f"Generated {event_count} events to '{output_path}'.")


if __name__ == "__main__":
    main()
